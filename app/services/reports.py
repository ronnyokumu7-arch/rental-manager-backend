import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from app.models.bookings import Booking, BookingStatus
from app.models.clients import Client
from app.models.invoices import Invoice, InvoiceStatus
from app.models.payments import Payment
from app.models.subscriptions import Subscription, SubscriptionStatus
from app.models.tenants import Tenant
from app.models.vehicles import Vehicle, VehicleStatus


BRAND = colors.HexColor("#1a1a2e")
ACCENT = colors.HexColor("#4f8cff")
LIGHT = colors.HexColor("#f7f7f7")
MUTED = colors.HexColor("#888888")


# ---------------------------------------------------------------------------
# Data builders
# ---------------------------------------------------------------------------

def get_revenue_summary(
    db: Session,
    tenant_id: Optional[int],
    start_date: Optional[datetime],
    end_date: Optional[datetime],
) -> dict:
    query = db.query(Booking).filter(
        Booking.status == BookingStatus.completed,
    )
    if tenant_id:
        query = query.filter(Booking.tenant_id == tenant_id)
    if start_date:
        query = query.filter(Booking.start_date >= start_date)
    if end_date:
        query = query.filter(Booking.end_date <= end_date)

    bookings = query.all()
    total_revenue = sum(b.total_amount for b in bookings)
    total_bookings = len(bookings)

    by_vehicle = {}
    for b in bookings:
        v = db.query(Vehicle).filter(Vehicle.id == b.vehicle_id).first()
        label = f"{v.make} {v.model} ({v.plate_number})" if v else f"Vehicle {b.vehicle_id}"
        by_vehicle[label] = by_vehicle.get(label, 0) + b.total_amount

    by_client = {}
    for b in bookings:
        c = db.query(Client).filter(Client.id == b.client_id).first()
        label = c.full_name if c else f"Client {b.client_id}"
        by_client[label] = by_client.get(label, 0) + b.total_amount

    return {
        "total_revenue": total_revenue,
        "total_bookings": total_bookings,
        "average_booking_value": round(total_revenue / total_bookings, 2) if total_bookings else 0,
        "by_vehicle": sorted(by_vehicle.items(), key=lambda x: x[1], reverse=True),
        "by_client": sorted(by_client.items(), key=lambda x: x[1], reverse=True),
        "currency": "KES",
    }


def get_booking_summary(
    db: Session,
    tenant_id: Optional[int],
    start_date: Optional[datetime],
    end_date: Optional[datetime],
) -> dict:
    query = db.query(Booking)
    if tenant_id:
        query = query.filter(Booking.tenant_id == tenant_id)
    if start_date:
        query = query.filter(Booking.start_date >= start_date)
    if end_date:
        query = query.filter(Booking.end_date <= end_date)

    bookings = query.all()
    by_status = {}
    for s in BookingStatus:
        by_status[s.value] = sum(1 for b in bookings if b.status == s)

    return {
        "total": len(bookings),
        "by_status": by_status,
        "archived": sum(1 for b in bookings if b.is_archived),
    }


def get_vehicle_utilisation(
    db: Session,
    tenant_id: int,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
) -> list:
    vehicles = db.query(Vehicle).filter(Vehicle.tenant_id == tenant_id).all()
    results = []

    for v in vehicles:
        query = db.query(Booking).filter(
            Booking.vehicle_id == v.id,
            Booking.status.in_([BookingStatus.completed, BookingStatus.active]),
        )
        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)

        bookings = query.all()
        days_rented = sum((b.end_date - b.start_date).days for b in bookings)
        total_revenue = sum(b.total_amount for b in bookings)

        if start_date and end_date:
            period_days = (end_date - start_date).days or 1
        else:
            period_days = 30

        utilisation_pct = round((days_rented / period_days) * 100, 1)

        results.append({
            "vehicle": f"{v.make} {v.model} ({v.year})",
            "plate": v.plate_number,
            "status": v.status.value,
            "days_rented": days_rented,
            "utilisation_pct": min(utilisation_pct, 100),
            "total_revenue": total_revenue,
            "total_bookings": len(bookings),
        })

    return sorted(results, key=lambda x: x["utilisation_pct"], reverse=True)


def get_client_activity(
    db: Session,
    tenant_id: int,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
) -> list:
    clients = db.query(Client).filter(Client.tenant_id == tenant_id).all()
    results = []

    for c in clients:
        query = db.query(Booking).filter(Booking.client_id == c.id)
        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)

        bookings = query.all()
        completed = [b for b in bookings if b.status == BookingStatus.completed]
        total_spent = sum(b.total_amount for b in completed)

        results.append({
            "client": c.full_name,
            "phone": c.phone,
            "status": c.status.value,
            "total_bookings": len(bookings),
            "completed_bookings": len(completed),
            "total_spent": total_spent,
        })

    return sorted(results, key=lambda x: x["total_spent"], reverse=True)


def get_overdue_bookings(db: Session, tenant_id: int) -> list:
    today = datetime.now(timezone.utc).date()
    bookings = db.query(Booking).filter(
        Booking.tenant_id == tenant_id,
        Booking.status == BookingStatus.active,
        Booking.end_date < today,
    ).all()

    results = []
    for b in bookings:
        client = db.query(Client).filter(Client.id == b.client_id).first()
        vehicle = db.query(Vehicle).filter(Vehicle.id == b.vehicle_id).first()
        days_overdue = (today - b.end_date).days
        results.append({
            "booking_id": b.id,
            "client": client.full_name if client else "—",
            "phone": client.phone if client else "—",
            "vehicle": f"{vehicle.make} {vehicle.model}" if vehicle else "—",
            "plate": vehicle.plate_number if vehicle else "—",
            "end_date": str(b.end_date),
            "days_overdue": days_overdue,
            "total_amount": b.total_amount,
        })

    return sorted(results, key=lambda x: x["days_overdue"], reverse=True)


def get_platform_revenue(db: Session) -> dict:
    tenants = db.query(Tenant).all()
    results = []
    grand_total = 0

    for t in tenants:
        bookings = db.query(Booking).filter(
            Booking.tenant_id == t.id,
            Booking.status == BookingStatus.completed,
        ).all()
        revenue = sum(b.total_amount for b in bookings)
        grand_total += revenue
        results.append({
            "tenant": t.name,
            "plan": t.plan,
            "subscription_status": t.subscription_status.value,
            "total_bookings": len(bookings),
            "total_revenue": revenue,
        })

    return {
        "tenants": sorted(results, key=lambda x: x["total_revenue"], reverse=True),
        "grand_total": grand_total,
        "total_tenants": len(tenants),
    }


def get_subscription_health(db: Session) -> dict:
    by_status = {}
    for s in SubscriptionStatus:
        count = db.query(Tenant).filter(
            Tenant.subscription_status == s,
        ).count()
        by_status[s.value] = count

    return {
        "by_status": by_status,
        "total_tenants": db.query(Tenant).count(),
    }


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------

def _report_header(elements, title: str, subtitle: str, styles):
    elements.append(Paragraph(title, ParagraphStyle(
        "RTitle", parent=styles["Heading1"], fontSize=20,
        textColor=BRAND, spaceAfter=2,
    )))
    elements.append(Paragraph(subtitle, ParagraphStyle(
        "RSub", parent=styles["Normal"], fontSize=10,
        textColor=MUTED, spaceAfter=12,
    )))
    elements.append(Spacer(1, 4 * mm))


def _styled_table(data, col_widths):
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def build_revenue_pdf(data: dict, tenant_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    _report_header(elements, "Revenue Report", tenant_name, styles)

    summary = [
        ["Metric", "Value"],
        ["Total revenue", f"KES {data['total_revenue']:,}"],
        ["Total bookings", str(data["total_bookings"])],
        ["Average booking value", f"KES {data['average_booking_value']:,}"],
    ]
    elements.append(_styled_table(summary, [100*mm, 70*mm]))
    elements.append(Spacer(1, 6*mm))

    elements.append(Paragraph("Revenue by Vehicle", ParagraphStyle(
        "SH", parent=styles["Heading2"], fontSize=12, textColor=BRAND, spaceAfter=4,
    )))
    vehicle_data = [["Vehicle", "Revenue (KES)"]] + [
        [v, f"{r:,}"] for v, r in data["by_vehicle"]
    ] or [["No data", "—"]]
    elements.append(_styled_table(vehicle_data, [120*mm, 50*mm]))
    elements.append(Spacer(1, 6*mm))

    elements.append(Paragraph("Revenue by Client", ParagraphStyle(
        "SH2", parent=styles["Heading2"], fontSize=12, textColor=BRAND, spaceAfter=4,
    )))
    client_data = [["Client", "Revenue (KES)"]] + [
        [c, f"{r:,}"] for c, r in data["by_client"]
    ] or [["No data", "—"]]
    elements.append(_styled_table(client_data, [120*mm, 50*mm]))

    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC",
        ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=MUTED),
    ))

    doc.build(elements)
    return buffer.getvalue()


def build_vehicle_utilisation_pdf(data: list, tenant_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    _report_header(elements, "Vehicle Utilisation Report", tenant_name, styles)

    rows = [["Vehicle", "Plate", "Status", "Days Rented", "Utilisation %", "Bookings", "Revenue (KES)"]]
    for r in data:
        rows.append([
            r["vehicle"], r["plate"], r["status"].title(),
            str(r["days_rented"]), f"{r['utilisation_pct']}%",
            str(r["total_bookings"]), f"{r['total_revenue']:,}",
        ])
    if len(rows) == 1:
        rows.append(["No data", "—", "—", "—", "—", "—", "—"])

    elements.append(_styled_table(rows, [55*mm, 30*mm, 25*mm, 25*mm, 25*mm, 20*mm, 35*mm]))
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC",
        ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=MUTED),
    ))

    doc.build(elements)
    return buffer.getvalue()


def build_overdue_pdf(data: list, tenant_name: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    _report_header(elements, "Overdue Bookings Report", tenant_name, styles)

    rows = [["Booking", "Client", "Phone", "Vehicle", "Plate", "Due Date", "Days Overdue", "Amount (KES)"]]
    for r in data:
        rows.append([
            f"#{r['booking_id']}", r["client"], r["phone"],
            r["vehicle"], r["plate"], r["end_date"],
            str(r["days_overdue"]), f"{r['total_amount']:,}",
        ])
    if len(rows) == 1:
        rows.append(["No overdue bookings", "—", "—", "—", "—", "—", "—", "—"])

    elements.append(_styled_table(rows, [20*mm, 35*mm, 28*mm, 35*mm, 25*mm, 25*mm, 22*mm, 25*mm]))
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC",
        ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, textColor=MUTED),
    ))

    doc.build(elements)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_excel_report(report_type: str, data, tenant_name: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header_row(row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    def style_data_row(row_num, num_cols, alt=False):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            if alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = thin

    ws.title = report_type.replace("_", " ").title()
    ws.append([f"{report_type.replace('_', ' ').title()} — {tenant_name}"])
    ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC"])
    ws.append([])

    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=13, color="1a1a2e")

    if report_type == "revenue":
        headers = ["Metric", "Value"]
        ws.append(headers)
        style_header_row(4, len(headers))
        rows = [
            ["Total Revenue", f"KES {data['total_revenue']:,}"],
            ["Total Bookings", data["total_bookings"]],
            ["Average Booking Value", f"KES {data['average_booking_value']:,}"],
        ]
        for i, row in enumerate(rows):
            ws.append(row)
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)

        ws.append([])
        ws.append(["Revenue by Vehicle", ""])
        style_header_row(ws.max_row, 2)
        for i, (v, r) in enumerate(data["by_vehicle"]):
            ws.append([v, f"KES {r:,}"])
            style_data_row(ws.max_row, 2, alt=i % 2 == 1)

        ws.append([])
        ws.append(["Revenue by Client", ""])
        style_header_row(ws.max_row, 2)
        for i, (c, r) in enumerate(data["by_client"]):
            ws.append([c, f"KES {r:,}"])
            style_data_row(ws.max_row, 2, alt=i % 2 == 1)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    elif report_type == "vehicle_utilisation":
        headers = ["Vehicle", "Plate", "Status", "Days Rented", "Utilisation %", "Bookings", "Revenue (KES)"]
        ws.append(headers)
        style_header_row(4, len(headers))
        for i, r in enumerate(data):
            ws.append([
                r["vehicle"], r["plate"], r["status"].title(),
                r["days_rented"], f"{r['utilisation_pct']}%",
                r["total_bookings"], r["total_revenue"],
            ])
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)
        for col, width in zip("ABCDEFG", [35, 15, 15, 15, 15, 12, 18]):
            ws.column_dimensions[col].width = width

    elif report_type == "client_activity":
        headers = ["Client", "Phone", "Status", "Total Bookings", "Completed", "Total Spent (KES)"]
        ws.append(headers)
        style_header_row(4, len(headers))
        for i, r in enumerate(data):
            ws.append([
                r["client"], r["phone"], r["status"].title(),
                r["total_bookings"], r["completed_bookings"], r["total_spent"],
            ])
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)
        for col, width in zip("ABCDEF", [30, 18, 15, 15, 15, 20]):
            ws.column_dimensions[col].width = width

    elif report_type == "overdue":
        headers = ["Booking ID", "Client", "Phone", "Vehicle", "Plate", "Due Date", "Days Overdue", "Amount (KES)"]
        ws.append(headers)
        style_header_row(4, len(headers))
        for i, r in enumerate(data):
            ws.append([
                f"#{r['booking_id']}", r["client"], r["phone"],
                r["vehicle"], r["plate"], r["end_date"],
                r["days_overdue"], r["total_amount"],
            ])
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)
        for col, width in zip("ABCDEFGH", [12, 25, 18, 25, 15, 15, 15, 18]):
            ws.column_dimensions[col].width = width

    elif report_type == "booking_summary":
        headers = ["Status", "Count"]
        ws.append(headers)
        style_header_row(4, len(headers))
        for i, (s, c) in enumerate(data["by_status"].items()):
            ws.append([s.title(), c])
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)
        ws.append([])
        ws.append(["Total", data["total"]])
        ws.append(["Archived", data["archived"]])
        for col, width in zip("AB", [25, 15]):
            ws.column_dimensions[col].width = width

    elif report_type == "platform_revenue":
        headers = ["Tenant", "Plan", "Subscription Status", "Total Bookings", "Revenue (KES)"]
        ws.append(headers)
        style_header_row(4, len(headers))
        for i, r in enumerate(data["tenants"]):
            ws.append([
                r["tenant"], r["plan"], r["subscription_status"].title(),
                r["total_bookings"], r["total_revenue"],
            ])
            style_data_row(5 + i, len(headers), alt=i % 2 == 1)
        ws.append([])
        ws.append(["Grand Total", "", "", "", f"KES {data['grand_total']:,}"])
        for col, width in zip("ABCDE", [30, 18, 22, 18, 20]):
            ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()